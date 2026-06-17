"""Add WBES variable mappings for priority B-READY Enterprise Surveys rows.

The script reads the reviewer-edited B-READY Enterprise Surveys question
inventory workbook, keeps rows with Priority == yes, and creates/replaces a
`wbes_variable_mapping` sheet with one row per candidate WBES microdata
variable.
"""

from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd


REVIEW_WORKBOOK = Path("Data/B-Ready/Raw/2025/bready_enterprise_survey_questions.xlsx")
WBES_DATASET = Path("Data/World Bank Enterprise Survey/New_Comprehensive_July_21_2025.dta")
QUESTION_SHEET = "enterprise_survey_questions"
MAPPING_SHEET = "wbes_variable_mapping"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

OUTPUT_COLUMNS = [
    "topic_sheet",
    "technical_variable_name",
    "priority",
    "bready_question",
    "bready_cameroon_response",
    "wbes_variable_order",
    "wbes_variable_name",
    "wbes_variable_label",
    "all_wbes_variable_names",
    "wbes_dataset_variable_exists",
    "construction_or_filter_note",
    "questionnaire_trace",
    "mapping_confidence",
    "mapping_status",
]


WBES_MAPPING = {
    ("02_Business_Location", "reg12"): {
        "wbes_variable_names": "g30a",
        "wbes_variable_labels": "g30a: How Much Of An Obstacle: Access To Land?",
        "construction_or_filter_note": "Obstacle/perception index built from access-to-land severity responses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.20: Land g30a is marked opinion based or sensitive.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("02_Business_Location", "bus3"): {
        "wbes_variable_names": "g2; g3",
        "wbes_variable_labels": "g2: application for construction-related permit submitted; g3: days to obtain construction-related permit",
        "construction_or_filter_note": "Median days among establishments that applied for a construction-related permit.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.20: g3 is days between completed/submitted application and permit granted.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("08_Dispute_Resolution", "disp6"): {
        "wbes_variable_names": "h30",
        "wbes_variable_labels": "h30: How Much Of An Obstacle: Courts",
        "construction_or_filter_note": "Obstacle/perception index built from courts severity responses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.24: h30 is listed with opinion-based or sensitive business-government relation items.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("08_Dispute_Resolution", "disp3"): {
        "wbes_variable_names": "j31",
        "wbes_variable_labels": "j31: In resolving commercial disputes, courts are independent and impartial",
        "construction_or_filter_note": "Perception share/index based on court independence/impartiality responses.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text, likely a newer module item.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("05_Financial_Services", "fin27"): {
        "wbes_variable_names": "k32",
        "wbes_variable_labels": "k32: Days to receive a decision about most recent loan application",
        "construction_or_filter_note": "Average/summary days among firms with a recent loan application decision.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("05_Financial_Services", "fin26"): {
        "wbes_variable_names": "k16; k17",
        "wbes_variable_labels": "k16: applied for new loans/lines of credit; k17: main reason for not applying for new loans/lines of credit",
        "construction_or_filter_note": "Percent reporting unfavorable rates, collateral, or procedures as main reason for not applying; requires coding selected k17 reasons and denominator among non-applicants.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.23: k17 is main reason to not apply for any line of credit or loan.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("05_Financial_Services", "fin30"): {
        "wbes_variable_names": "k36",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Cost per transaction to receive the main type of e-payment.",
        "questionnaire_trace": "Matched from WBES .dta variable label.",
        "mapping_confidence": "high",
        "mapping_status": "mapped_from_dta_label",
    },
    ("05_Financial_Services", "fin31"): {
        "wbes_variable_names": "k38",
        "wbes_variable_labels": "k38: Percentage of payments made using e-payments",
        "construction_or_filter_note": "Average proportion of payments made electronically.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("05_Financial_Services", "fin33"): {
        "wbes_variable_names": "k30",
        "wbes_variable_labels": "k30: How Much Of An Obstacle: Access To Finance",
        "construction_or_filter_note": "Obstacle/perception index built from access-to-finance severity responses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.23: k30 is marked opinion based or sensitive.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("06_International_Trade", "tr26"): {
        "wbes_variable_names": "d30b",
        "wbes_variable_labels": "d30b: How Much Of An Obstacle: Customs And Trade Regulations?",
        "construction_or_filter_note": "Obstacle/perception index built from customs and trade regulation severity responses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.14: d30a/d30b are opinion based or sensitive.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("06_International_Trade", "tr18_u"): {
        "wbes_variable_names": "d33a; d33b",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Days to clear all border control agencies; d33b records hours when clearance took less than one day and needs conversion to days for the constructed B-READY indicator.",
        "questionnaire_trace": "Matched from WBES .dta variable labels for exported goods border-control clearance.",
        "mapping_confidence": "high",
        "mapping_status": "mapped_from_dta_label",
    },
    ("06_International_Trade", "tr20"): {
        "wbes_variable_names": "d342; d34",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Costs to comply with all export requirements as percent of value exported; d342 is the combined/general cost variable and d34 is the FCA-specific source item.",
        "questionnaire_trace": "Matched from WBES .dta variable labels for export compliance costs.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("06_International_Trade", "tr24_u"): {
        "wbes_variable_names": "d40a; d40b",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Days to clear all border control agencies for imported goods; d40b records hours when clearance took less than one day and needs conversion to days for the constructed B-READY indicator.",
        "questionnaire_trace": "Matched from WBES .dta variable labels for imported goods border-control clearance.",
        "mapping_confidence": "high",
        "mapping_status": "mapped_from_dta_label",
    },
    ("06_International_Trade", "tr25"): {
        "wbes_variable_names": "d412; d41",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Costs to comply with all import requirements as percent of value imported; d412 is the combined/general cost variable and d41 is the DAP-specific source item.",
        "questionnaire_trace": "Matched from WBES .dta variable labels for import compliance costs.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("06_International_Trade", "in23"): {
        "wbes_variable_names": "d30a",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Obstacle/perception index built from transportation severity responses.",
        "questionnaire_trace": "Matched from WBES .dta variable label.",
        "mapping_confidence": "high",
        "mapping_status": "mapped_from_dta_label",
    },
    ("04_Labor", "wk28"): {
        "wbes_variable_names": "l30a",
        "wbes_variable_labels": "l30a: How Much Of An Obstacle: Labor Regulations?",
        "construction_or_filter_note": "Obstacle/perception index built from labor regulation severity responses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.27: l30a/l30b are opinion based or sensitive.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "comp2"): {
        "wbes_variable_names": "e1; e2b_ESBR",
        "wbes_variable_labels": "e1: main market for main product; e2b_ESBR: competitors for main product/service in main market, combined version",
        "construction_or_filter_note": "Percent reporting fewer than two competitors; exclude firms whose main market is international using e1.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.16: e2b is number of competitors in the establishment's market.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "comp1"): {
        "wbes_variable_names": "e1; e31a; e31b",
        "wbes_variable_labels": "",
        "construction_or_filter_note": "Market share of largest competitor; exclude firms whose main market is international using e1. Both e31a and e31b are present in the WBES extract and should be reviewed when constructing the final indicator.",
        "questionnaire_trace": "Matched from WBES .dta variable labels.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("09_Market_Competition", "comp5"): {
        "wbes_variable_names": "e1; e33",
        "wbes_variable_labels": "e1: main market for main product; e33: establishment can increase prices more than competitors without losing customers",
        "construction_or_filter_note": "Percent that cannot increase prices more than competitors without losing customers; exclude firms whose main market is international using e1.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("09_Market_Competition", "t11_4"): {
        "wbes_variable_names": "h5; sector_MS",
        "wbes_variable_labels": "h5: introduced new/significantly improved process; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among services firms; filter service firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.18: h5 defines new or improved processes.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "t11_3"): {
        "wbes_variable_names": "h5; sector_MS",
        "wbes_variable_labels": "h5: introduced new/significantly improved process; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among manufacturing firms; filter manufacturing firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.18: h5 defines new or improved processes.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "t11_5"): {
        "wbes_variable_names": "h8; sector_MS",
        "wbes_variable_labels": "h8: spent on R&D in last fiscal year; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among manufacturing firms; filter manufacturing firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.19: h8 defines spending on R&D.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "t11_6"): {
        "wbes_variable_names": "h8; sector_MS",
        "wbes_variable_labels": "h8: spent on R&D in last fiscal year; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among services firms; filter service firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.19: h8 defines spending on R&D.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "t1_1"): {
        "wbes_variable_names": "b8; sector_MS",
        "wbes_variable_labels": "b8: internationally-recognized quality certification; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among manufacturing firms; filter manufacturing firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.10: b8 defines internationally recognized quality certification.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "t1_2"): {
        "wbes_variable_names": "b8; sector_MS",
        "wbes_variable_labels": "b8: internationally-recognized quality certification; sector_MS: manufacturing or services",
        "construction_or_filter_note": "Percent among services firms; filter service firms with sector_MS.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.10: b8 defines internationally recognized quality certification.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("09_Market_Competition", "comp9"): {
        "wbes_variable_names": "j41",
        "wbes_variable_labels": "j41: Degree of difficulty to comply with govt contract tender requirements",
        "construction_or_filter_note": "Perception index among firms responding to government tender requirement difficulty.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("09_Market_Competition", "gend7"): {
        "wbes_variable_names": "j42; b4; b7a",
        "wbes_variable_labels": "j42: held a government contract in last 3 years; b4: any female owners; b7a: top manager is female",
        "construction_or_filter_note": "Percent female-owned or female-managed among firms that held a government contract in last 3 years.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf pp.9-10: b4 identifies female ownership and b7a identifies female top manager; j42 from WBES .dta label.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("07_Taxation", "tax1"): {
        "wbes_variable_names": "j35a",
        "wbes_variable_labels": "j35a: Total annual time spent on tax compliance (hours)",
        "construction_or_filter_note": "Median annual hours spent on tax compliance.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("07_Taxation", "tax2"): {
        "wbes_variable_names": "j36",
        "wbes_variable_labels": "j36: In last FY, did establishment file taxes electronically",
        "construction_or_filter_note": "Percent of firms filing taxes electronically.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("07_Taxation", "tax6"): {
        "wbes_variable_names": "j38; j40",
        "wbes_variable_labels": "j38: applied for VAT refund in last 3 years; j40: main reason for not applying for VAT tax refund",
        "construction_or_filter_note": "Percent reporting too long/complicated refund process among firms eligible for a VAT refund; requires coding j40 reason categories and denominator.",
        "questionnaire_trace": "WBES .dta variable labels; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
    ("03_Utility_Services", "in1"): {
        "wbes_variable_names": "c3; c4",
        "wbes_variable_labels": "c3: applied for electrical connection; c4: days to receive electrical connection service",
        "construction_or_filter_note": "Median days among establishments that applied for an electrical connection.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.10: c4 is days between completed/submitted application and connection provision.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("03_Utility_Services", "in4"): {
        "wbes_variable_names": "c6; c9a",
        "wbes_variable_labels": "c6: experienced power outages; c9a: losses as percent of annual sales due to power outages",
        "construction_or_filter_note": "Median outage losses as percent of annual sales, generally among firms with outages/nonmissing losses.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.11: c9a/c9b record losses because of power outages.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("03_Utility_Services", "in2"): {
        "wbes_variable_names": "c6; c7",
        "wbes_variable_labels": "c6: experienced power outages; c7: number of power outages in a typical month",
        "construction_or_filter_note": "Average number of electrical outages in a typical month.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.11: c7 is number of outages in a typical month.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("03_Utility_Services", "in18"): {
        "wbes_variable_names": "c15",
        "wbes_variable_labels": "c15: Did you experience insufficient water supply for production in last fiscal year?",
        "construction_or_filter_note": "Percent of firms experiencing water insufficiencies.",
        "questionnaire_trace": "ES_QuestionnaireManual_2019.pdf p.11: c15 defines insufficient water supply.",
        "mapping_confidence": "high",
        "mapping_status": "mapped",
    },
    ("03_Utility_Services", "in21"): {
        "wbes_variable_names": "c39",
        "wbes_variable_labels": "c39: Did you experience internet disruptions in last fiscal year?",
        "construction_or_filter_note": "Percent of firms experiencing internet disruptions.",
        "questionnaire_trace": "WBES .dta variable label; not found in the available 2019 questionnaire manual text.",
        "mapping_confidence": "medium",
        "mapping_status": "mapped_from_dta_label",
    },
}


def find_repo_root() -> Path:
    here = Path.cwd().resolve()
    for candidate in [here, *here.parents]:
        if (candidate / "AGENTS.md").exists():
            return candidate
    raise FileNotFoundError("Run this script from inside the CEMAC repository.")


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def yes(value: object) -> bool:
    return str(value or "").strip().lower() == "yes"


def split_variable_names(mapping: dict[str, str]) -> list[str]:
    names = str(mapping["wbes_variable_names"]).split(";")
    return [name.strip() for name in names if name.strip()]


def split_variable_labels(mapping: dict[str, str], names: list[str]) -> dict[str, str]:
    labels = {}
    raw_labels = str(mapping["wbes_variable_labels"])
    for piece in raw_labels.split(";"):
        if ":" not in piece:
            continue
        name, label = piece.split(":", 1)
        labels[name.strip()] = label.strip()

    return {name: labels.get(name, "") for name in names}


def read_wbes_variable_labels(dataset_path: Path) -> dict[str, str]:
    reader = pd.io.stata.StataReader(dataset_path)
    return reader.variable_labels()


def read_priority_rows(workbook, wbes_variable_labels: dict[str, str]) -> list[dict[str, object]]:
    worksheet = workbook[QUESTION_SHEET]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {normalize_header(header): idx for idx, header in enumerate(headers)}

    required = [
        "topic_sheet",
        "technical_variable_name",
        "survey_question_text",
        "cameroon_response",
        "priority",
    ]
    missing = [header for header in required if header not in header_index]
    if missing:
        raise KeyError(f"Missing required column(s) in {QUESTION_SHEET}: {missing}")

    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not yes(row[header_index["priority"]]):
            continue

        topic_sheet = str(row[header_index["topic_sheet"]])
        technical_name = str(row[header_index["technical_variable_name"]])
        key = (topic_sheet, technical_name)
        mapping = WBES_MAPPING.get(key)
        if mapping is None:
            mapping = {
                "wbes_variable_names": "",
                "wbes_variable_labels": "",
                "construction_or_filter_note": "",
                "questionnaire_trace": "",
                "mapping_confidence": "none",
                "mapping_status": "needs_mapping",
            }

        variable_names = split_variable_names(mapping)
        all_variable_names = "; ".join(variable_names)

        if not variable_names:
            variable_names = [""]

        for variable_order, variable_name in enumerate(variable_names, start=1):
            variable_exists = variable_name in wbes_variable_labels
            rows.append(
                {
                    "topic_sheet": topic_sheet,
                    "technical_variable_name": technical_name,
                    "priority": row[header_index["priority"]],
                    "bready_question": row[header_index["survey_question_text"]],
                    "bready_cameroon_response": row[header_index["cameroon_response"]],
                    "wbes_variable_order": variable_order,
                    "wbes_variable_name": variable_name,
                    "wbes_variable_label": wbes_variable_labels.get(variable_name, ""),
                    "all_wbes_variable_names": all_variable_names,
                    "wbes_dataset_variable_exists": "yes" if variable_exists else "no",
                    "construction_or_filter_note": mapping["construction_or_filter_note"],
                    "questionnaire_trace": mapping["questionnaire_trace"],
                    "mapping_confidence": mapping["mapping_confidence"],
                    "mapping_status": mapping["mapping_status"],
                }
            )

    return rows


def write_mapping_sheet(workbook, rows: list[dict[str, object]]) -> None:
    if MAPPING_SHEET in workbook.sheetnames:
        del workbook[MAPPING_SHEET]

    worksheet = workbook.create_sheet(MAPPING_SHEET)
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 24,
        "B": 24,
        "C": 12,
        "D": 64,
        "E": 20,
        "F": 18,
        "G": 20,
        "H": 58,
        "I": 30,
        "J": 18,
        "K": 62,
        "L": 62,
        "M": 18,
        "N": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def validate(rows: list[dict[str, object]]) -> None:
    indicator_keys = {
        (row["topic_sheet"], row["technical_variable_name"]) for row in rows
    }
    if not indicator_keys:
        raise AssertionError("No priority indicators found.")

    missing = [row for row in rows if row["mapping_status"] == "needs_mapping"]
    if missing:
        missing_keys = [
            f"{row['topic_sheet']}:{row['technical_variable_name']}" for row in missing
        ]
        raise AssertionError(f"Missing WBES mappings: {missing_keys}")

    nonexistent = [
        row for row in rows
        if row["wbes_variable_name"] and row["wbes_dataset_variable_exists"] != "yes"
    ]
    if nonexistent:
        bad_names = [
            f"{row['topic_sheet']}:{row['technical_variable_name']}->{row['wbes_variable_name']}"
            for row in nonexistent
        ]
        raise AssertionError(f"Mapped variables not found in WBES dataset: {bad_names}")

    expected = {
        ("02_Business_Location", "reg12"): "g30a",
        ("07_Taxation", "tax1"): "j35a",
        ("06_International_Trade", "tr18_u"): "d33a",
        ("05_Financial_Services", "fin30"): "k36",
    }
    by_key = {
        (row["topic_sheet"], row["technical_variable_name"], row["wbes_variable_order"]): row
        for row in rows
    }
    for key, expected_vars in expected.items():
        if (key[0], key[1], 1) not in by_key:
            continue
        actual = by_key[(key[0], key[1], 1)]["wbes_variable_name"]
        if actual != expected_vars:
            raise AssertionError(f"Unexpected mapping for {key}: {actual}")


def main() -> int:
    repo_root = find_repo_root()
    workbook_path = repo_root / REVIEW_WORKBOOK
    wbes_dataset_path = repo_root / WBES_DATASET
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if not wbes_dataset_path.exists():
        raise FileNotFoundError(wbes_dataset_path)

    workbook = load_workbook(workbook_path)
    if QUESTION_SHEET not in workbook.sheetnames:
        raise KeyError(f"Workbook is missing {QUESTION_SHEET}.")

    wbes_variable_labels = read_wbes_variable_labels(wbes_dataset_path)
    rows = read_priority_rows(workbook, wbes_variable_labels)
    validate(rows)
    write_mapping_sheet(workbook, rows)
    workbook.save(workbook_path)

    status_counts = {}
    for row in rows:
        status = row["mapping_status"]
        status_counts[status] = status_counts.get(status, 0) + 1

    print(f"Updated {workbook_path}")
    indicator_count = len(
        {(row["topic_sheet"], row["technical_variable_name"]) for row in rows}
    )
    print(f"Priority indicators mapped: {indicator_count}")
    print(f"Priority variable rows written: {len(rows)}")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
