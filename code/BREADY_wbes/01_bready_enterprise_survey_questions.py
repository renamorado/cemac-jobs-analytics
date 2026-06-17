"""Extract B-READY Enterprise Surveys questions with score hierarchy labels.

This script reads the official B-READY 2025 EconomyAnswer workbook, keeps
questions whose data source is Enterprise Surveys, and exports one tidy review
workbook. It also parses the topic-score workbook's header hierarchy and joins
labels through an explicit topic/technical-variable crosswalk.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ECONOMY_ANSWER = Path("Data/B-Ready/Raw/2025/02_B-READY-2025-EconomyAnswer.xlsx")
TOPIC_SCORES = Path("Data/B-Ready/Raw/2025/01_B-READY-2025-PILLAR-TOPIC-SCORES.xlsx")
OUTPUT_FILE = Path("output/tables/bready_enterprise_survey_questions.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOPIC_PREFIX_RE = re.compile(r"^\d+_")


SCORE_LABEL_CROSSWALK = {
    ("02_Business_Location", "bus3"): "Time to Obtain a Construction-Related Permit",
    ("02_Business_Location", "reg12"): "Major Constraints on Access to Land",
    ("03_Utility_Services", "in1"): "Time to Obtain an Electricity Connection",
    ("03_Utility_Services", "in2"): "Reliability of Electricity Supply",
    ("03_Utility_Services", "in3"): "Reliability of Electricity Supply",
    ("03_Utility_Services", "in4"): "Reliability of Electricity Supply",
    ("03_Utility_Services", "in5"): "Time to Obtain a Water Connection",
    ("03_Utility_Services", "in9"): "Reliability of Electricity Supply",
    ("03_Utility_Services", "in18"): "Reliability of Water Supply",
    ("03_Utility_Services", "in21"): "Reliability of Internet Supply",
    ("03_Utility_Services", "in22"): "Time to Obtain an Internet Connection",
    ("04_Labor", "reg10"): "Percent of firms visited or inspected for workplace health and safety",
    ("04_Labor", "reg11"): "Percent of firms with a report issued by inspectorate for workplace health and safety",
    ("04_Labor", "tax7"): "Ratio of social contributions",
    ("04_Labor", "wk1"): "Percent of firms with formal training programs for its permanent, full-time work",
    ("04_Labor", "wk24"): "Weeks to dismiss full-time permanent worker",
    ("04_Labor", "wk25"): "Weeks paid in severance",
    ("04_Labor", "wk26"): "Percent of firms involved in labor dispute over last 3 years",
    ("04_Labor", "wk27"): "Months to resolve labor dispute",
    ("04_Labor", "wk28"): "Perception index of labor regulations as a constraint",
    ("05_Financial_Services", "fin26"): "Obstacles to Obtaining a Loan",
    ("05_Financial_Services", "fin27"): "Time to Obtain a Loan",
    ("05_Financial_Services", "fin28"): "Usage level of Receiving e-Payments",
    ("05_Financial_Services", "fin29"): "Time to Receive e-Payments",
    ("05_Financial_Services", "fin30"): "Cost of Receiving e-Payments",
    ("05_Financial_Services", "fin31"): "Usage level of Making e-Payments",
    ("05_Financial_Services", "fin32"): "Cost of Making e-Payments",
    ("06_International_Trade", "in23"): "Share of Firms Identifying Transportation as Major or Severe Constraints",
    ("06_International_Trade", "tr18_u"): "Total Time to Comply with Export Requirements",
    ("06_International_Trade", "tr19"): "Share of Trading Firms Exporting Digitally Ordered Goods",
    ("06_International_Trade", "tr20"): "Total Cost to Comply with Export Requirements",
    ("06_International_Trade", "tr24_u"): "Total Time to Comply with Import Requirements",
    ("06_International_Trade", "tr25"): "Total Cost to Comply with Import Requirements",
    ("06_International_Trade", "tr26"): "Share of Firms Identifying Customs and Trade Regulations as Major or Severe Constraints",
    ("07_Taxation", "tax1"): "Total Time for Preparation, Filing and Payment",
    ("07_Taxation", "tax2"): "The percentage of Firms Filing and Paying Taxes Electronically",
    ("07_Taxation", "tax3"): "The percentage of Firms Filing and Paying Taxes Electronically",
    ("07_Taxation", "tax4"): "Total Time Needed to Complete the Audit",
    ("07_Taxation", "tax5"): "Obtaining a VAT Refund in Practice",
    ("07_Taxation", "tax6"): "Obtaining a VAT Refund in Practice",
    ("07_Taxation", "tax7"): "Effective Tax Rate (ETR) for Employment-Based Taxes and Social Contributions",
    ("07_Taxation", "tax8"): "Effective Tax Rate (ETR) for Profit Taxes",
    ("08_Dispute_Resolution", "disp3"): "In resolving commercial disputes, courts are independent and impartial",
    ("08_Dispute_Resolution", "disp4"): "Arbitration is a reliable mechanism to resolve commercial disputes",
    ("08_Dispute_Resolution", "disp5"): "Mediation is a reliable mechanism to resolve commercial disputes",
    ("08_Dispute_Resolution", "disp6"): "Courts are not an obstacle to business operations",
    ("09_Market_Competition", "comp1"): "Market Concentration (Market Share of Largest Competitor)",
    ("09_Market_Competition", "comp2"): "Market Structure (Number of Firms that Compete in the Market)",
    ("09_Market_Competition", "comp3"): "Market Structure (Number of Firms that Compete in the Market)",
    ("09_Market_Competition", "comp4"): "Market Structure (Number of Firms that Compete in the Market)",
    ("09_Market_Competition", "comp5"): "Pricing Power (Ability to Change Prices without Losing Costumers)",
    ("09_Market_Competition", "comp6"): "Government Intervention in Prices",
    ("09_Market_Competition", "comp7"): "Changes in the Level of Competition",
    ("09_Market_Competition", "comp8"): "Easiness to Switch Internet Provider",
    ("09_Market_Competition", "comp9"): "Firms' Perceptions on the Difficulty to Meet the Administrative Requirements to Participate in Tenders",
    ("09_Market_Competition", "gend7"): "Gender Gap in Government Suppliers",
    ("09_Market_Competition", "reg9"): "Time to Receive a Payment from a Government Contract",
    ("09_Market_Competition", "t1_1"): "Use of International Quality Certifications (Manufacturing)",
    ("09_Market_Competition", "t1_2"): "Use of International Quality Certifications (Services)",
    ("09_Market_Competition", "t11_1"): "Proportion of Firms with Product Innovation (Manufacturing)",
    ("09_Market_Competition", "t11_2"): "Proportion of Firms with Product Innovation (Services)",
    ("09_Market_Competition", "t11_3"): "Proportion of Firms with Process Innovation (Manufacturing)",
    ("09_Market_Competition", "t11_4"): "Proportion of Firms with Process Innovation (Services)",
    ("09_Market_Competition", "t11_5"): "Proportion of Firms that Spent on R&D (Manufacturing)",
    ("09_Market_Competition", "t11_6"): "Proportion of Firms that Spent on R&D (Services)",
}


OUTPUT_COLUMNS = [
    "topic_sheet",
    "topic_name",
    "pillar_label",
    "category_label",
    "subcategory_label",
    "score_indicator_label",
    "technical_variable_name",
    "survey_question_text",
    "data_source",
    "cameroon_response",
    "economy_count",
    "nonmissing_response_count",
    "hierarchy_match_status",
]


def normalize_text(value: object) -> str:
    """Normalize labels for exact, punctuation-insensitive lookup."""
    text = "" if value is None else str(value)
    text = text.replace("\u2019", "'").replace("\u2018", "'")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = text.replace("\u00a0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text).strip().lower()
    text = re.sub(r"[^a-z0-9&]+", "", text)
    return text


def clean_topic_name(sheet_name: str) -> str:
    return TOPIC_PREFIX_RE.sub("", sheet_name).replace("_", " ")


def is_nonmissing_response(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return False
        if stripped.lower() in {"no data", "not available", "n/a"}:
            return False
    if isinstance(value, (int, float)) and value == -777:
        return False
    return True


def find_repo_root() -> Path:
    here = Path.cwd().resolve()
    for candidate in [here, *here.parents]:
        if (candidate / "AGENTS.md").exists():
            return candidate
    raise FileNotFoundError("Run this script from inside the CEMAC repository.")


def parse_score_hierarchy(score_path: Path) -> dict[tuple[str, str], dict[str, str]]:
    workbook = load_workbook(score_path, read_only=True, data_only=True)
    hierarchy: dict[tuple[str, str], dict[str, str]] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in {"ABOUT", "00_B-READY_Pillar_Score"}:
            continue

        worksheet = workbook[sheet_name]
        header_rows = list(worksheet.iter_rows(min_row=1, max_row=4, values_only=True))
        for column_index in range(1, worksheet.max_column + 1):
            pillar, category, subcategory, score_label = [
                header_rows[row_index][column_index - 1] for row_index in range(4)
            ]

            if not score_label or column_index <= 3:
                continue

            key = (sheet_name, normalize_text(score_label))
            hierarchy[key] = {
                "pillar_label": "" if pillar is None else str(pillar),
                "category_label": "" if category is None else str(category),
                "subcategory_label": "" if subcategory is None else str(subcategory),
                "score_indicator_label": str(score_label),
            }

    workbook.close()
    return hierarchy


def read_enterprise_survey_questions(economy_answer_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(economy_answer_path, read_only=True, data_only=True)
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    scanned_topic_count = 0

    for sheet_name in workbook.sheetnames:
        if sheet_name == "README":
            continue

        scanned_topic_count += 1
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            economy, technical_name, question_text, response, data_source = row[:5]
            if data_source != "Enterprise Surveys":
                continue
            if technical_name is None:
                raise ValueError(f"Missing technical variable name in {sheet_name}.")

            key = (sheet_name, str(technical_name))
            entry = grouped.setdefault(
                key,
                {
                    "topic_sheet": sheet_name,
                    "topic_name": clean_topic_name(sheet_name),
                    "technical_variable_name": str(technical_name),
                    "survey_question_text_values": set(),
                    "data_source_values": set(),
                    "cameroon_response": None,
                    "economies": set(),
                    "nonmissing_response_count": 0,
                },
            )

            entry["survey_question_text_values"].add("" if question_text is None else str(question_text))
            entry["data_source_values"].add(str(data_source))
            entry["economies"].add(str(economy))
            if is_nonmissing_response(response):
                entry["nonmissing_response_count"] += 1
            if economy == "Cameroon":
                entry["cameroon_response"] = response

    workbook.close()

    if scanned_topic_count != 10:
        raise AssertionError(f"Expected 10 topic sheets, scanned {scanned_topic_count}.")

    rows = []
    for key in sorted(grouped):
        entry = grouped[key]
        question_values = sorted(entry.pop("survey_question_text_values"))
        source_values = sorted(entry.pop("data_source_values"))
        rows.append(
            {
                **entry,
                "survey_question_text": " | ".join(question_values),
                "data_source": " | ".join(source_values),
                "economy_count": len(entry.pop("economies")),
            }
        )

    return rows


def enrich_with_hierarchy(
    rows: list[dict[str, object]],
    hierarchy: dict[tuple[str, str], dict[str, str]],
) -> list[dict[str, object]]:
    duplicate_check = set()
    enriched = []

    for row in rows:
        unique_key = (row["topic_sheet"], row["technical_variable_name"])
        if unique_key in duplicate_check:
            raise AssertionError(f"Duplicate exported key: {unique_key}")
        duplicate_check.add(unique_key)

        target_label = SCORE_LABEL_CROSSWALK.get(unique_key)
        if target_label is None:
            hierarchy_values = {
                "pillar_label": "",
                "category_label": "",
                "subcategory_label": "",
                "score_indicator_label": "",
            }
            match_status = "needs_review"
        else:
            lookup_key = (row["topic_sheet"], normalize_text(target_label))
            hierarchy_values = hierarchy.get(lookup_key)
            if hierarchy_values is None:
                hierarchy_values = {
                    "pillar_label": "",
                    "category_label": "",
                    "subcategory_label": "",
                    "score_indicator_label": target_label,
                }
                match_status = "not_in_score_workbook"
            else:
                match_status = "matched"

        enriched.append({**row, **hierarchy_values, "hierarchy_match_status": match_status})

    return enriched


def write_inventory_workbook(rows: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "enterprise_survey_questions"
    worksheet.append(OUTPUT_COLUMNS)

    for row in rows:
        worksheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 24,
        "B": 26,
        "C": 42,
        "D": 38,
        "E": 44,
        "F": 48,
        "G": 24,
        "H": 70,
        "I": 20,
        "J": 20,
        "K": 16,
        "L": 22,
        "M": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.save(output_path)


def validate_output(rows: list[dict[str, object]]) -> None:
    if len(rows) != 66:
        raise AssertionError(f"Expected 66 Enterprise Surveys questions, found {len(rows)}.")

    if any(row["data_source"] != "Enterprise Surveys" for row in rows):
        raise AssertionError("Export contains a non-Enterprise Surveys source.")

    keys = [(row["topic_sheet"], row["technical_variable_name"]) for row in rows]
    if len(keys) != len(set(keys)):
        raise AssertionError("Duplicate topic_sheet + technical_variable_name rows found.")

    samples = {
        ("03_Utility_Services", "in1"): 21,
        ("06_International_Trade", "tr18_u"): 5.249497526,
        ("07_Taxation", "tax1"): 42,
        ("09_Market_Competition", "comp1"): 89.92508605,
    }
    by_key = {(row["topic_sheet"], row["technical_variable_name"]): row for row in rows}
    for key, expected in samples.items():
        actual = by_key[key]["cameroon_response"]
        if isinstance(expected, float):
            if abs(float(actual) - expected) > 1e-8:
                raise AssertionError(f"Cameroon response mismatch for {key}: {actual}")
        elif actual != expected:
            raise AssertionError(f"Cameroon response mismatch for {key}: {actual}")

    matched_rows = [row for row in rows if row["hierarchy_match_status"] == "matched"]
    if not matched_rows:
        raise AssertionError("No rows matched to score hierarchy.")

    for row in matched_rows:
        for column in ["pillar_label", "category_label", "subcategory_label"]:
            if row[column] == "":
                raise AssertionError(
                    f"Matched row has missing {column}: "
                    f"{row['topic_sheet']} {row['technical_variable_name']}"
                )


def print_summary(rows: list[dict[str, object]], output_path: Path) -> None:
    status_counts: defaultdict[str, int] = defaultdict(int)
    topic_counts: defaultdict[str, int] = defaultdict(int)

    for row in rows:
        status_counts[str(row["hierarchy_match_status"])] += 1
        topic_counts[str(row["topic_sheet"])] += 1

    print(f"Exported {len(rows)} Enterprise Surveys question rows to {output_path}.")
    print("Hierarchy match status:")
    for status in sorted(status_counts):
        print(f"  {status}: {status_counts[status]}")
    print("Topic counts:")
    for topic in sorted(topic_counts):
        print(f"  {topic}: {topic_counts[topic]}")


def main() -> int:
    repo_root = find_repo_root()
    economy_answer_path = repo_root / ECONOMY_ANSWER
    score_path = repo_root / TOPIC_SCORES
    output_path = repo_root / OUTPUT_FILE

    for path in [economy_answer_path, score_path]:
        if not path.exists():
            raise FileNotFoundError(path)

    hierarchy = parse_score_hierarchy(score_path)
    rows = read_enterprise_survey_questions(economy_answer_path)
    rows = enrich_with_hierarchy(rows, hierarchy)
    rows = [{column: row.get(column, "") for column in OUTPUT_COLUMNS} for row in rows]
    validate_output(rows)
    write_inventory_workbook(rows, output_path)
    print_summary(rows, output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
